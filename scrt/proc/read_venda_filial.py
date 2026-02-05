import os
import json
import time
import logging
import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
import gspread
import xlrd
from openpyxl import Workbook, load_workbook

# Configure logging for GitHub Actions
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

class VendasFilialError(Exception):
    """Custom exception for Vendas Filial processing failures"""
    pass

class GoogleSheetsError(Exception):
    """Custom exception for Google Sheets operations"""
    pass

class VendasFilialProcessor:
    def __init__(self, directory: str = "."):
        self.directory = directory
        logging.info(f"Initialized Vendas Filial Processor with directory: {directory}")
    
    def get_latest_file(self):
        """Get the most recently modified Excel file (.xls or .xlsx)"""
        try:
            import glob
            # Search for both .xls and .xlsx files
            xls_files = glob.glob(os.path.join(self.directory, '*.xls'))
            xlsx_files = glob.glob(os.path.join(self.directory, '*.xlsx'))
            all_files = xls_files + xlsx_files
            
            if not all_files:
                raise VendasFilialError(f"No Excel files found in {self.directory}")
            
            latest_file = max(all_files, key=os.path.getmtime)
            logging.info(f"Found latest file: {os.path.basename(latest_file)}")
            return latest_file
            
        except Exception as e:
            raise VendasFilialError(f"Failed to find latest file: {str(e)}")
    
    def convert_xls_to_xlsx(self, file_path: str) -> str:
        """Convert .xls file to .xlsx format if needed"""
        if file_path.lower().endswith(".xlsx"):
            logging.info("File already .xlsx, skipping conversion.")
            return file_path
        
        logging.info("Converting .xls to .xlsx...")
        
        try:
            # Try to open as real .xls
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            
            # Create new .xlsx workbook
            wb = Workbook()
            ws = wb.active
            
            # Copy all rows
            for row_idx in range(sheet.nrows):
                ws.append(sheet.row_values(row_idx))
            
            # Save as .xlsx
            new_path = file_path.replace(".xls", ".xlsx")
            wb.save(new_path)
            
            logging.info(f"✓ Converted to: {os.path.basename(new_path)}")
            return new_path
            
        except xlrd.biffh.XLRDError:
            # File is not a real .xls (might be .xls extension but Excel 2007+ format)
            logging.warning("File is not a real .xls. Renaming to .xlsx.")
            new_path = file_path.replace(".xls", ".xlsx")
            os.rename(file_path, new_path)
            return new_path
        except Exception as e:
            raise VendasFilialError(f"Failed to convert file: {str(e)}")
    
    def _get_column_index(self, sheet, header_name: str) -> int:
        """Find column index by header name (case-insensitive)"""
        header_name = header_name.strip().lower()
        
        for col in sheet.iter_cols(1, sheet.max_column):
            cell_value = str(col[0].value).strip().lower() if col[0].value else ""
            if cell_value == header_name:
                return col[0].column - 1  # Convert to 0-based index
        
        raise VendasFilialError(f"Header '{header_name}' not found in sheet")
    
    def process_excel_data(self, file_path: str):
        """Process Vendas Filial Excel file"""
        if not os.path.exists(file_path):
            raise VendasFilialError(f"File does not exist: {file_path}")
        
        if os.path.getsize(file_path) == 0:
            raise VendasFilialError(f"File is empty: {file_path}")
        
        try:
            logging.info(f"Processing Vendas Filial Excel file: {os.path.basename(file_path)}")
            
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # Log first few rows for debugging
            logging.info("First 5 rows (for debugging):")
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                logging.info(f"  Row {i}: {row}")
            
            # Find required columns
            try:
                col_codigo = self._get_column_index(sheet, "código")
                col_total_vlr_venda = self._get_column_index(sheet, "total vlr. venda")
                col_total_vlr_custo = self._get_column_index(sheet, "total vlr. custo")
                col_vlr_descto = self._get_column_index(sheet, "vlr. descto")
                col_ticket_medio = self._get_column_index(sheet, "ticket médio venda/devol.")
                
                logging.info("✓ Found all required columns")
                
            except VendasFilialError as e:
                # Try alternative column names
                logging.warning(f"Standard column not found: {str(e)}")
                
                # Fallback: try to find columns with partial matches
                col_codigo = self._get_column_index(sheet, "código")
                col_total_vlr_venda = self._get_column_index(sheet, "total vlr")
                col_total_vlr_custo = self._get_column_index(sheet, "custo")
                col_vlr_descto = self._get_column_index(sheet, "descto")
                col_ticket_medio = self._get_column_index(sheet, "ticket")
            
            # Process rows
            data = []
            filial_number = None
            faturamento_hb = None
            
            for row in sheet.iter_rows(values_only=True):
                codigo = str(row[col_codigo]).strip() if row[col_codigo] is not None else ""
                
                # Detect Filial row
                if codigo.lower() == "filial:":
                    next_cell = row[col_codigo + 1] if (col_codigo + 1) < len(row) else None
                    if next_cell:
                        filial_number = str(next_cell).split()[0]
                        logging.info(f"Found Filial: {filial_number}")
                    continue
                
                # Detect Faturamento HB row (código 8000)
                if codigo == "8000":
                    faturamento_hb = row[col_total_vlr_venda]
                    continue
                
                # Detect "Total Filial" rows
                if codigo.lower().startswith("total filial"):
                    if not filial_number:
                        logging.warning("Total Filial found without Filial number")
                        continue
                    
                    # Extract values
                    custo_total = row[col_total_vlr_custo] if col_total_vlr_custo < len(row) else None
                    faturamento_total = row[col_vlr_descto] if col_vlr_descto < len(row) else None
                    ticket_medio = row[col_ticket_medio] if col_ticket_medio < len(row) else None
                    
                    data.append({
                        "Filial": filial_number.zfill(2),  # Format as 2-digit
                        "Faturamento HB": faturamento_hb,
                        "Custo Total": custo_total,
                        "Faturamento Total": faturamento_total,
                        "Ticket Médio": ticket_medio,
                    })
                    
                    logging.info(f"Processed Filial {filial_number}: "
                                f"HB={faturamento_hb}, Custo={custo_total}, "
                                f"Faturamento={faturamento_total}, Ticket={ticket_medio}")
                    
                    # Reset for next filial
                    filial_number = None
                    faturamento_hb = None
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            if df.empty:
                raise VendasFilialError("No valid filial data found after processing")
            
            # Ensure proper data types
            numeric_cols = ["Faturamento HB", "Custo Total", "Faturamento Total", "Ticket Médio"]
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            logging.info(f"✓ Processing complete. Rows processed: {len(df)}")
            return df
            
        except Exception as e:
            raise VendasFilialError(f"Failed to process Excel file: {str(e)}")


class VendasFilialSheetsUploader:
    def __init__(self, credentials_json: str, sheet_id: str):
        if not credentials_json:
            raise GoogleSheetsError("Google credentials JSON cannot be empty")
        if not sheet_id:
            raise GoogleSheetsError("Sheet ID cannot be empty")
        
        self.credentials_json = credentials_json
        self.sheet_id = sheet_id
        self.client = None
        
    def authenticate(self):
        """Authenticate with Google Sheets API"""
        try:
            logging.info("Authenticating with Google Sheets API...")
            creds_dict = json.loads(self.credentials_json)
            scope = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"
            ]
            creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
            self.client = gspread.authorize(creds)
            logging.info("✓ Authentication successful")
            
        except json.JSONDecodeError as e:
            raise GoogleSheetsError(f"Invalid JSON in credentials: {str(e)}")
        except Exception as e:
            raise GoogleSheetsError(f"Authentication failed: {str(e)}")
    
    def retry_api_call(self, func, retries=3, delay=2):
        """Retry API calls on 500 errors"""
        for i in range(retries):
            try:
                return func()
            except HttpError as error:
                if hasattr(error, "resp") and error.resp.status == 500:
                    logging.warning(f"APIError 500 encountered. Retrying {i + 1}/{retries}...")
                    time.sleep(delay)
                else:
                    raise
        raise GoogleSheetsError("Max retries reached for API call")
    
    def upload_vendas_filial(self, df: pd.DataFrame, worksheet_name: str = "VENDAS_FILIAL"):
        """Upload Vendas Filial data to Google Sheets"""
        if not self.client:
            raise GoogleSheetsError("Client not authenticated. Call authenticate() first")
        
        if df.empty:
            raise GoogleSheetsError("DataFrame is empty. Nothing to upload")
        
        try:
            # Open spreadsheet and worksheet
            logging.info(f"Opening spreadsheet with ID: {self.sheet_id}")
            spreadsheet = self.client.open_by_key(self.sheet_id)
            worksheet = spreadsheet.worksheet(worksheet_name)
            logging.info(f"✓ Accessed worksheet: {worksheet_name}")
            
            # Prepare data
            logging.info("Preparing data for Google Sheets...")
            df = df.fillna("")  # Ensure no NaN values
            
            # Clear existing data (keep header row)
            logging.info("Clearing existing data (keeping headers)...")
            worksheet.batch_clear(["A2:Z"])  # Clear from row 2 onward
            
            # Update data starting from A1 (headers + data)
            data = [df.columns.tolist()] + df.values.tolist()
            
            logging.info(f"Uploading {len(df)} rows of data...")
            update_func = lambda: worksheet.update(
                "A1",
                data,
                value_input_option="USER_ENTERED"
            )
            self.retry_api_call(update_func)
            
            logging.info("✓ Vendas Filial data uploaded successfully to Google Sheets")
            
        except gspread.exceptions.APIError as e:
            raise GoogleSheetsError(f"Google Sheets API error: {str(e)}")
        except Exception as e:
            raise GoogleSheetsError(f"Failed to upload Vendas Filial data: {str(e)}")


def main():
    """Main execution with proper error handling for GitHub Actions"""
    try:
        # Get environment variables
        download_dir = os.getenv("DOWNLOAD_DIR", "/home/runner/work/comp_meta/comp_meta/")
        sheet_id = os.getenv("SOURCE_SHEET_ID")
        gsa_credentials = os.getenv("GSA_CREDENTIALS")
        
        # Validate environment variables
        if not sheet_id:
            raise GoogleSheetsError("SHEET_ID environment variable not set")
        if not gsa_credentials:
            raise GoogleSheetsError("GSA_CREDENTIALS environment variable not set")
        
        logging.info("Environment variables loaded successfully")
        logging.info(f"Download directory: {download_dir}")
        logging.info(f"Directory exists: {os.path.exists(download_dir)}")
        
        # Wait for potential file downloads to complete
        logging.info("Waiting for potential file downloads to complete...")
        time.sleep(10)
        
        # Process Excel file
        processor = VendasFilialProcessor(directory=download_dir)
        latest_file = processor.get_latest_file()
        
        # Convert .xls to .xlsx if needed
        if latest_file.lower().endswith(".xls"):
            latest_file = processor.convert_xls_to_xlsx(latest_file)
        
        # Process the file
        processed_df = processor.process_excel_data(latest_file)
        
        # Show preview
        logging.info("\n=== VENDAS FILIAL DATA PREVIEW ===")
        logging.info(f"Total filials processed: {len(processed_df)}")
        logging.info(f"Columns: {list(processed_df.columns)}")
        if not processed_df.empty:
            logging.info("\nData preview:")
            logging.info(processed_df.to_string())
        
        # Upload to Google Sheets
        uploader = VendasFilialSheetsUploader(gsa_credentials, sheet_id)
        uploader.authenticate()
        uploader.upload_vendas_filial(processed_df, "VENDAS_FILIAL")
        
        # Clean up: remove the processed file
        logging.info(f"Removing processed file: {os.path.basename(latest_file)}")
        os.remove(latest_file)
        
        logging.info("✓ Process completed successfully")
        return 0  # Success exit code
        
    except (VendasFilialError, GoogleSheetsError) as e:
        logging.error(f"✗ {e.__class__.__name__}: {str(e)}")
        return 1  # Business logic failure
    except Exception as e:
        logging.error(f"✗ Unexpected error: {str(e)}")
        return 2  # Unexpected failure


if __name__ == "__main__":
    # Exit with proper code for GitHub Actions
    exit_code = main()
    exit(exit_code)
